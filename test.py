# encoding=utf8
from selenium import webdriver
from selenium.common.exceptions import ElementNotVisibleException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
import time
import win32gui
import openpyxl
import pyautogui


# 前台开启浏览器模式
def openChrome():
    # 加启动配置
    option = webdriver.ChromeOptions()
    option.add_argument('disable-infobars')


    # 打开chrome浏览器
    driver = webdriver.Chrome(chrome_options=option)
    return driver

# 授权操作
def operationAuth(driver):
    url = "http://192.168.1.64/"
    driver.get(url)

    try:
        elem_pw = driver.find_elements_by_xpath("//input[@type='password']")
        for elem in elem_pw:
            elem.send_keys("Buchou123")
        # class="aui_state_highlight" class="aui_close"  placeholder="用户名"
        driver.find_element_by_xpath("//button[@class='aui_state_highlight']").click()
        driver.find_element_by_xpath("//a[class='aui_close']").click()
    except ElementNotVisibleException:
        print("已注册")
        driver.find_element_by_xpath("//input[@type='password']").clear()
    driver.find_element_by_xpath("//input[@type='password']").send_keys("Buchou123")
    driver.find_element_by_xpath("//input[@placeholder='用户名']").send_keys("admin")
    driver.find_element_by_xpath("//button").click()
    print("登陆成功")
    time.sleep(2)
    driver.find_element_by_xpath("//a[@ng-bind='oLan.config']").click()
    advanced_setting(driver)
    osd(driver)
    region_cover(driver)
    record_project(driver)
    storage_manage(driver)
    video_setting(driver)
    record_device_num(driver)
    yunmou(driver)
def advanced_setting(driver):

    driver.find_element_by_xpath("//span[@class='menu-icon network-icon']").click()
    driver.find_element_by_xpath("//span[@ng-bind='oMenuLan.advancedSettings']").click()
    time.sleep(1)
    print("选中高级设置")
    driver.find_element_by_xpath("//a[@id = 'ui-id-14']").click()
    time.sleep(1)
    driver.find_element_by_xpath("//select[@ng-model='oParams.szAccessType']").click()
    driver.find_element_by_xpath("//option[@value='ezviz']").click()
    time.sleep(1)
    if(driver.find_element_by_xpath("//input[@ng-model='oParams.bEnableEzviz']").is_selected()!=True):
        driver.find_element_by_xpath("//input[@ng-model='oParams.bEnableEzviz']").click()
        time.sleep(2)
        driver.find_elements_by_xpath("//input")[0].send_keys('12341234q')
        driver.find_elements_by_xpath("//input")[1].send_keys('12341234q')
        driver.find_element_by_xpath("//button[@class='aui_state_highlight']").click()
        driver.find_element_by_xpath("//button[@class='btn btn-primary btn-save']").click()
    print("萤石云配置成功")

#osd配置
def osd(driver):
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.image"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//a[@ng-bind="oImageLan.OSDSettings"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@ng-model="oOsdParams.cameraName"]').clear()
    book = openpyxl.load_workbook('养户摄像头对应表.xlsx')
    sheet = book.worksheets[0]

    for column in sheet.columns:
        for i in range(len(column)):
            if (column[i].value == None):
                devices_num = i
                print("已录入设备数目为" + str(devices_num))
                break
    for column in sheet.columns:
        name = column[devices_num].value
        break
    print("当前录入养户为" + name)
    book.close()
    driver.find_element_by_xpath('//input[@ng-model="oOsdParams.cameraName"]').send_keys(name)
    driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
    print("osd修改完成")

#遮挡区域配置
def region_cover(driver):
    driver.find_element_by_xpath('//span[@class="menu-icon event-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//a[@ng-bind="oLan.videoTamper"]').click()
    time.sleep(1)
    # print("选中图像")
    if(driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').is_selected()!=True):
        driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').click()
        driver.find_element_by_xpath('//button[@ng-click="draw()"]').click()
        time.sleep(1)
        hwnd_title = {}

        def get_all_hwnd(hwnd, mouse):
            if (win32gui.IsWindow(hwnd) and
                    win32gui.IsWindowEnabled(hwnd) and
                    win32gui.IsWindowVisible(hwnd)):
                hwnd_title.update({hwnd: win32gui.GetWindowText(hwnd)})

        win32gui.EnumWindows(get_all_hwnd, 0)



        for h, t in hwnd_title.items():
            if t:
                # print(h, t)
                if t == 'WindowControl':
                    left, top, right, bottom = win32gui.GetWindowRect(h)
                    # print(left, top, right, bottom)

                    pyautogui.moveTo(left +1, top+1 )
                    time.sleep(0.5)
                    pyautogui.dragTo(x=right-1, y=bottom-1,duration=0.1)
                    time.sleep(0.5)
        slider = driver.find_element_by_xpath('//div[@class="slider"]')
        print(slider.size['width'])
        action = ActionChains(driver)
        action.move_to_element(slider)
        action.move_by_offset(slider.size['width']/2,0)
        action.click()
        action.perform()
        driver.find_element_by_xpath('//button[@class="btn btn-primary btn-save"]').click()
        print("绘制遮挡完成")
    driver.find_element_by_xpath('//div[@class ="step last"]').click()
    time.sleep(1)
    if(driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').is_selected()!=True):
        driver.find_element_by_xpath('//input[@id="enableVideoTamper"]').click()
    if(driver.find_element_by_xpath('//input[@ng-click="alarmLinkAll()"]').is_selected()!=True):
        driver.find_element_by_xpath('//input[@ng-click="alarmLinkAll()"]').click()
    if(driver.find_element_by_xpath('//input[@ng-click="normalLinkAll()"]').is_selected()!=True):
        driver.find_element_by_xpath('//input[@ng-click="normalLinkAll()"]').click()
    driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
    print("联动方式设置完成")

def record_project(driver):
    driver.find_element_by_xpath('//span[@class ="menu-icon storage-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.scheduleSettings"]').click()
    time.sleep(1)
    if(driver.find_element_by_xpath('//input[@class="checkbox ng-pristine ng-valid"]').is_selected()!=True):
        driver.find_element_by_xpath('//input[@class="checkbox ng-pristine ng-valid"]').click()
        driver.find_element_by_xpath('//span[@id="recordDiv_deleteAll_txt"]').click()

        sliders = driver.find_elements_by_xpath('//div[@class="timeplan_daytimeplan"]')
        for slider in sliders:
            # print(slider.size)
            action = ActionChains(driver)
            action.move_to_element(slider)
            action.move_by_offset(-slider.size['width']/2,0)
            action.click_and_hold()
            action.move_by_offset(slider.size['width'] , 0)
            action.release()
            # action.drag_and_drop_by_offset(slider,slider.size['width']/2,0)
            action.perform()
        driver.find_element_by_xpath('//button[@ng-click="save()"]').click()
    print("录像计划配置完成")
def video_setting(driver):
    driver.find_element_by_xpath('//span[@class="menu-icon video-icon"]').click()
    time.sleep(1)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.resolution"]')).select_by_index(0)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.bitrateType"]')).select_by_index(1)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.videoQuality"]')).select_by_index(2)
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.frameRate"]')).select_by_value("11")
    driver.find_element_by_xpath('//input[@ng-model="oVideoParams.maxBitrate"]').clear()
    driver.find_element_by_xpath('//input[@ng-model="oVideoParams.maxBitrate"]').send_keys("2048")
    Select(driver.find_element_by_xpath('//select[@ng-model="oVideoParams.videoEncoding"]')).select_by_index(1)
    driver.find_element_by_xpath('//button[@class="btn btn-primary btn-save"]').click()

def storage_manage(driver):
    #存储配置

    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.storageManagement"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@class="checkbox"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//input[@ng-value="oLan.format"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//button[@class="aui_state_highlight"]').click()
    driver.find_element_by_xpath('//input[@ng-model="oParams.iVideoQuotaRatio"]').clear()
    driver.find_element_by_xpath('//input[@ng-model="oParams.iVideoQuotaRatio"]').send_keys("90")
    while(True):
        try:
            driver.find_element_by_xpath('//img[@src="../ui/images/artDialog/loading.gif"]')
        except:
            break

    driver.find_element_by_xpath('//span[@ng-bind="oLan.save"]').click()
    print("存储配置完成")

def record_device_num(driver):
    driver.find_element_by_xpath('//span[@class="menu-icon system-icon"]').click()
    time.sleep(1)
    driver.find_element_by_xpath('//span[@ng-bind="oMenuLan.systemSettings"]').click()
    time.sleep(1)
    device_num=driver.find_element_by_xpath('//input[@ng-model = "oSettingBasicInfo.szSerialNo"]').get_attribute('value')
    print("序列号为"+device_num)
    try:
        book = openpyxl.load_workbook('养户摄像头对应表.xlsx')
        sheet = book.worksheets[0]
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows):
            if (sheet.cell(row=i, column=2).value == None):
                sheet.cell(row=i, column=2, value=device_num)
                break
        book.save('养户摄像头对应表.xlsx')
        book.close()
    except:
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")
        print("录入设备序列号失败，把excel关闭，不然读写不了")


def yunmou(driver):
    driver.close()
    # 加启动配置
    option = webdriver.ChromeOptions()
    option.add_argument('disable-infobars')

    # 打开chrome浏览器
    driver = webdriver.Chrome(chrome_options=option)
    url ="https://www.hik-cloud.com/chain/login/index.html#/login"
    driver.get(url)


# 方法主入口
if __name__ == '__main__':
    # 加启动配置
    driver = openChrome()
    print('''
    请确认以下事项:
    1.同一内网只能连一台摄像头，如果公司其他人也在连着同一wifi配置摄像头，请切换wifi或者等对方配完再配
    2.内网更改摄像头信息时后台过程中不要移动鼠标直到配置完成
    3.配置过程中不要打开excel，否则会导致无法写入
    4.一定要用管理员权限打开本文件!!!!!!!!否则必定出错!!!!!!!!
    5.配置时保持chromedriver.exe打开
    确认无误后按任意键开始自动配置摄像头
    ''')
    input()
    operationAuth(driver)
    print('''
    ```````````````````````````*******``.............................```**`*`*******
``````````````````````````*`********`..........................````**```********
`````````````````````````****`*``**````````..................``*`***************
````````````*********```**```````````````````````````````````*******************
******************``*```````````````****``````````````*``***********************
******************````````````````*`*******```````````******************@*******
***************````````````````````******`````````````*````*@*******************
******```````````````````````````****````*``*`*`````````````**********@@********
******``````````````````````````****```````````````````````````*****@@@@****`***
*```````````````````````````````***`*````````````*``*````````````**@*@@*********
``````````````````````````````******```````````````***`````````````**********@*`
````````````````**@@*````````*`****```````````````****````````````````**********
````````....```****@@**```````*`*``````````````````**``**```````````````********
...```........`.@**@@@@``````````````````````````````**`**```````````````*******
.````........`*.***@@*````````````````````*@@@@@@*`*``*````````````````````***@*
..``.......````*****```````````````*****@@`.**@*@@@***``````````````````````****
.``........````````````````````````****@@```@@@@@@@@**``````````````````````****
.`.........````````````````````````************@@`*```````````````````````````**
...........```.`````````````````````*````*****`**```````````````````````````````
.............````````````````````````````````````````````..`````````````````````
`.........``*********``````.```````````````````````````````.````````````````````
`.`.....`@@*@@@@@@@@****``````````````````````````````````......````````````````
````....`@@@*@@@@@@@@@**`....`````````````````````````````.......```````````````
........****@@@@@@@@@@@*````````````````````````..``````````````````````````````
......```@**@@@@@@@@@***`````````````````````.```.`````````````````````````````*
....`````***@@@@@@*****`````````````````````````````.``````````````````````````*
....````*****@********``.````````````````````````````````````````````````````***
....````**************`````````````````````````````````````````````````````*****
....`````**@@@@@@*******`*```````````*``````````````````````````````````````****
....`.`.`******@@@@@@@@*```````````****``````````````````````````````````````***
......````***********@@@@@@`@@@@@*@***`````````````````````````````````````**`**
......```````************@**********```````````````````````````````````````*****
.......``````*`******```*****`*****``````````````````````````````````````*``****
.....````````````````````````````````````````````````````````````````````*******
...````````````````````````````````````````````````````````````````````*`*******
....```````````````````````````````````````````````````````````````````****`****
...`````````````````````````````````````````````````````````````````````**`*****
..``````````````````````````````````````````````````````````````````````********
.``````````````````````````````````````````````````````````````````*````********
..````````````````````````````````````````````````````````````````````**********
————————————————
安装完成，直接关了就完事了''')